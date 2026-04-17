import json
import re
import os
import sys
import time
import shutil
from pathlib import Path
from openpyxl import load_workbook

# 基于仓库根目录推导路径，避免目录迁移后脚本失效
REPO_ROOT = Path(__file__).resolve().parents[1]
DASHBOARD_DIR = REPO_ROOT / "dashboard"
DOCS_DIR = REPO_ROOT / "docs"

EXCEL_PATH_1 = REPO_ROOT / "生词表.xlsx"
EXCEL_PATH_2 = REPO_ROOT / "雅思题目.xlsx"
OUTPUT_JS = DASHBOARD_DIR / "script.js"
DOCS_JS = DOCS_DIR / "script.js"
SOURCE_INDEX = DASHBOARD_DIR / "index.html"
SOURCE_CSS = DASHBOARD_DIR / "styles.css"
SOURCE_VOCAB_JSON = DASHBOARD_DIR / "vocabulary.json"

def clean_text(text):
    if not text: return ""
    text = text.replace('\u200b', '').replace('\xa0', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def parse_single_excel(path, default_chapter_name, force_chapter=None):
    structured_data = {}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        current_chapter = default_chapter_name if not force_chapter else force_chapter

        for row in ws.iter_rows(min_row=1, values_only=True):
            val_a = clean_text(row[0] if len(row) > 0 else '')
            val_b = clean_text(row[1] if len(row) > 1 else '')

            if not val_a and not val_b:
                continue

            # For vocabulary list, detect chapters
            if not force_chapter and val_a and not val_b:
                if re.match(r'(chapter|第\s*\d+\s*章)', val_a, re.I):
                    current_chapter = val_a
                    continue

            # Skip date markers like "4.4日", "4月5日"
            if re.match(r'^\d+\.\d+日?$', val_a) or re.match(r'^\d+月\d+日?$', val_a):
                continue

            # Skip pure numeric marker rows
            if val_a.isdigit() and not val_b:
                continue

            entry = {'word': val_a, 'definition': val_b}
            if current_chapter not in structured_data:
                structured_data[current_chapter] = []
            structured_data[current_chapter].append(entry)
                        
    except Exception as e:
        print(f"[{time.strftime('%H:%M:%S')}] 读取 {os.path.basename(path)} 失败: {e}")
    return structured_data

def sync_data():
    print(f"[{time.strftime('%H:%M:%S')}] 正在同步单词表...")
    try:
        data1 = {}
        if os.path.exists(EXCEL_PATH_1):
            data1 = parse_single_excel(EXCEL_PATH_1, "精选词汇")
            
        data2 = {}
        if os.path.exists(EXCEL_PATH_2):
            data2 = parse_single_excel(EXCEL_PATH_2, "雅思强化题目生词", force_chapter="雅思强化题目生词")
            
        # Merge data
        all_data = {}
        for k, v in data1.items():
            all_data[k] = v
        for k, v in data2.items():
            if k in all_data:
                all_data[k].extend(v)
            else:
                all_data[k] = v
                
        # IELTS 强化词库按 Excel 倒序展示：表格最下面的词显示在最上面
        if "雅思强化题目生词" in all_data:
            all_data["雅思强化题目生词"].reverse()

        vocab_json = json.dumps(all_data, ensure_ascii=False)
        sync_time = time.strftime('%Y-%m-%d %H:%M:%S')

        with open(OUTPUT_JS, 'r', encoding='utf-8') as src:
            js_content = src.read()

        pattern = r'const allData = .*?;\nconst syncTime = ".*?";'
        replacement = f'const allData = {vocab_json};\nconst syncTime = "{sync_time}";'
        new_content, count = re.subn(pattern, replacement, js_content, flags=re.S)

        if count != 1:
            raise RuntimeError("未找到 allData/syncTime 区块，无法安全更新 script.js")

        with open(OUTPUT_JS, 'w', encoding='utf-8') as out:
            out.write(new_content)

        os.makedirs(DOCS_DIR, exist_ok=True)
        with open(DOCS_JS, 'w', encoding='utf-8') as out:
            out.write(new_content)
        shutil.copy2(SOURCE_INDEX, DOCS_DIR / 'index.html')
        shutil.copy2(SOURCE_CSS, DOCS_DIR / 'styles.css')
        if os.path.exists(SOURCE_VOCAB_JSON):
            shutil.copy2(SOURCE_VOCAB_JSON, DOCS_DIR / 'vocabulary.json')

        total_words = sum(len(v) for v in all_data.values())
        print(f"[{time.strftime('%H:%M:%S')}] 同步成功！更新了 {total_words} 个词条。")
    except Exception as e:
        print(f"[{time.strftime('%H:%M:%S')}] 同步时出错: {e}")

def print_startup_info():
    print(">>> 雅思词汇双轨同步助手已启动")
    print(f">>> 正在监控:")
    print(f"    1. {os.path.basename(EXCEL_PATH_1)}")
    print(f"    2. {os.path.basename(EXCEL_PATH_2)}")
    print(">>> 提示：每次你在 Excel 中点击保存，网页内容就会自动更新（刷新网页即可见）。")

def watch_and_sync():
    print_startup_info()
    last_mtime1 = 0
    last_mtime2 = 0
    while True:
        try:
            mtime1 = os.path.getmtime(EXCEL_PATH_1) if os.path.exists(EXCEL_PATH_1) else 0
            mtime2 = os.path.getmtime(EXCEL_PATH_2) if os.path.exists(EXCEL_PATH_2) else 0

            if mtime1 != last_mtime1 or mtime2 != last_mtime2:
                sync_data()
                last_mtime1 = mtime1
                last_mtime2 = mtime2
        except Exception:
            pass
        time.sleep(2)

if __name__ == "__main__":
    if "--once" in sys.argv:
        sync_data()
    else:
        watch_and_sync()
